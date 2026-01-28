#!/usr/bin/env python3
"""Aggregate per-case stem metrics XML files into a single Excel workbook."""

from __future__ import annotations

import argparse
import re
import sys
import xml.etree.ElementTree as ET
from collections import defaultdict
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook
except ImportError as exc:  # pragma: no cover - handled during runtime
    raise SystemExit(
        "openpyxl is required for Excel export. Install it via 'pip install -r requirements_excel.txt'."
    ) from exc


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Scan for *_stem_metrics.xml files (emitted by load_nifti_and_stem.py) and merge them into a multi-sheet "
            "Excel workbook with per-case details plus per-user summaries."
        )
    )
    parser.add_argument(
        "--root",
        required=True,
    )
    parser.add_argument(
        "--output",
        default="stem_metrics.xlsx",
        help="Destination .xlsx path (default: %(default)s).",
    )
    parser.add_argument(
        "--anteversion-excel",
        help=(
            "Optional path to a batchCompareStudies Excel file. When provided, the "
            "Femoral_Anteversion_Angles sheet and its anteversion_angle_deg column are "
            "joined into the GruenHU sheet by case id."
        ),
    )
    parser.add_argument(
        "--hu-heatmap",
        choices=("auto", "ezplan", "ezplan-2024"),
        default="auto",
        help=(
            "Select which HU heatmap XML to read for Gruen summaries. "
            "auto uses the default export if present, otherwise any available heatmap."
        ),
    )
    parser.add_argument(
        "--pattern",
        default="**/*_stem_metrics.xml",
        help="Glob pattern (relative to --root) used to discover per-case XML exports.",
    )
    parser.add_argument(
        "--quiet",
        action="store_true",
        help="Suppress per-file log lines while parsing XML files.",
    )
    return parser.parse_args()


def _collect_metric_files(root: Path, pattern: str) -> list[Path]:
    files = [path for path in root.glob(pattern) if path.is_file()]
    files.sort()
    return files


def _to_float(value: Any, default: float = 0.0) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _to_int(value: Any, default: int = 0) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _normalize_case_id(value: str) -> str:
    clean = value.replace("\\", "/").strip()
    clean = clean.strip("/")
    if not clean:
        return ""
    return clean.split("/")[-1]


def _extract_case_id_from_path(path_value: str) -> str:
    if not path_value:
        return ""
    parts = path_value.replace("\\", "/").split("/")
    pattern = re.compile(r"\d{3}-[A-Z]-\d{2,}")
    for part in parts:
        if pattern.fullmatch(part):
            return part
    return ""


def _heatmap_suffix(value: str | None) -> str:
    if not value or value == "auto":
        return ""
    return "_" + value.replace("-", "_")


def _select_gruen_xml(candidate_dir: Path, base_prefix: str, heatmap: str | None) -> Path | None:
    suffix = _heatmap_suffix(heatmap)
    if suffix:
        preferred = candidate_dir / f"{base_prefix}_stem_local_gruen_hu_summary{suffix}.xml"
        if preferred.is_file():
            return preferred
    default_path = candidate_dir / f"{base_prefix}_stem_local_gruen_hu_summary.xml"
    if default_path.is_file():
        return default_path
    pattern = f"{base_prefix}_stem_local_gruen_hu_summary*.xml"
    matches = sorted(candidate_dir.glob(pattern))
    if not matches:
        matches = sorted(candidate_dir.glob("*_stem_local_gruen_hu_summary*.xml"))
    if not matches:
        return None
    if suffix:
        for match in matches:
            if suffix in match.stem:
                return match
    return matches[0]


def _select_partition_xml(candidate_dir: Path, base_prefix: str, heatmap: str | None) -> Path | None:
    suffix = _heatmap_suffix(heatmap)
    if suffix:
        preferred = candidate_dir / f"{base_prefix}_hu_summary{suffix}.xml"
        if preferred.is_file():
            return preferred
    default_path = candidate_dir / f"{base_prefix}_hu_summary.xml"
    if default_path.is_file():
        return default_path
    pattern = f"{base_prefix}_hu_summary*.xml"
    matches = sorted(candidate_dir.glob(pattern))
    if not matches:
        matches = sorted(candidate_dir.glob("*_hu_summary*.xml"))
    if not matches:
        return None
    if suffix:
        for match in matches:
            if suffix in match.stem:
                return match
    return matches[0]


def _parse_metric_file(path: Path, hu_heatmap: str | None = None) -> dict[str, Any]:
    tree = ET.parse(path)
    root = tree.getroot()
    stem_elem = root.find("stem")
    stem_data = {
        "uid": stem_elem.get("uid") if stem_elem is not None else "",
        "manufacturer": stem_elem.get("manufacturer") if stem_elem is not None else "",
        "enumName": stem_elem.get("enumName") if stem_elem is not None else "",
        "friendlyName": stem_elem.get("friendlyName") if stem_elem is not None else "",
        "rotationMode": stem_elem.get("rotationMode") if stem_elem is not None else "",
        "requestedSide": stem_elem.get("requestedSide") if stem_elem is not None else "",
        "configuredSide": stem_elem.get("configuredSide") if stem_elem is not None else "",
        "state": stem_elem.get("state") if stem_elem is not None else "",
        "seedplanSource": stem_elem.get("seedplanSource") if stem_elem is not None else "",
        "rccId": stem_elem.get("rccId") if stem_elem is not None else "",
    }
    zones: dict[str, dict[str, float | int]] = {}
    for zone_elem in root.findall("./zones/zone"):
        name = zone_elem.get("name") or "Unknown"
        zones[name] = {
            "count": _to_int(zone_elem.get("count")),
            "percent": _to_float(zone_elem.get("percent")),
            "low": _to_float(zone_elem.get("low")),
            "high": _to_float(zone_elem.get("high")),
        }
    partition_summary: dict[str, dict[str, float | int]] = {}
    base_prefix = path.name.replace("_stem_metrics.xml", "")
    partition_xml = _select_partition_xml(path.parent, base_prefix, hu_heatmap)
    if partition_xml is not None and partition_xml.is_file():
        try:
            part_tree = ET.parse(partition_xml)
            part_root = part_tree.getroot()
            for zone_elem in part_root.findall("./zone"):
                zone_name = zone_elem.get("name") or zone_elem.get("id") or "Unknown"
                partition_summary[zone_name] = {}
                for tag_elem in zone_elem.findall("./tag"):
                    tag_name = tag_elem.get("name") or "Unknown"
                    partition_summary[zone_name][tag_name] = _to_float(tag_elem.get("percent"))
        except Exception:
            partition_summary = {}
    gruen_summary: dict[int, dict[str, tuple[int, float]]] = {}
    gruen_candidates: list[Path] = [path.parent]
    original_vtp_value = root.get("stemOriginalVtp")
    if original_vtp_value:
        gruen_candidates.append(Path(original_vtp_value).parent)
    screenshot_dir = root.get("screenshotDir")
    if screenshot_dir:
        gruen_candidates.append(Path(screenshot_dir))
    gruen_xml: Path | None = None
    gruen_xml_found = False
    for candidate_dir in gruen_candidates:
        if not candidate_dir:
            continue
        candidate_dir = Path(candidate_dir)
        candidate_path = _select_gruen_xml(candidate_dir, base_prefix, hu_heatmap)
        if candidate_path is not None and candidate_path.is_file():
            gruen_xml = candidate_path
            gruen_xml_found = True
            break
    if gruen_xml_found:
        try:
            gruen_tree = ET.parse(gruen_xml)
            gruen_root = gruen_tree.getroot()
            for zone_elem in gruen_root.findall("./zone"):
                zone_id = _to_int(zone_elem.get("id"))
                if zone_id <= 0:
                    continue
                gruen_summary[zone_id] = {}
                for tag_elem in zone_elem.findall("./tag"):
                    tag_name = tag_elem.get("name") or "Unknown"
                    gruen_summary[zone_id][tag_name] = (
                        _to_int(tag_elem.get("count")),
                        _to_float(tag_elem.get("percent")),
                    )
        except Exception:
            gruen_summary = {}
    data = {
        "path": path,
        "user_id": root.get("userId", ""),
        "case_id": root.get("caseId", ""),
        "config_label": root.get("stemConfigLabel", ""),
        "config_source": root.get("stemConfigSource", ""),
        "config_index": _to_int(root.get("stemConfigIndex"), default=-1),
        "hip_config_name": root.get("hipConfigName") or root.get("stemConfigLabel") or root.get("stemConfigSource") or "",
        "hip_config_description": root.get("hipConfigPrettyName") or root.get("hipConfigName") or "",
        "array": root.get("array", ""),
        "total_points": _to_int(root.get("totalPoints")),
        "scalar_min": _to_float(root.get("scalarMin")),
        "scalar_max": _to_float(root.get("scalarMax")),
        "volume_path": root.get("volumePath", ""),
        "seedplan_path": root.get("seedplanPath", ""),
        "screenshot_dir": root.get("screenshotDir", ""),
        "original_vtp": root.get("stemOriginalVtp", ""),
        "generated_at": root.get("generatedAt", ""),
        "stem": stem_data,
        "zones": zones,
        "partition_summary": partition_summary,
        "gruen_summary": gruen_summary,
        "gruen_xml_found": gruen_xml_found,
    }
    return data


def _autosize_columns(sheet) -> None:
    for column_cells in sheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            value = cell.value
            if value is None:
                continue
            max_length = max(max_length, len(str(value)))
        sheet.column_dimensions[column_letter].width = min(max_length + 2, 60)


def _load_anteversion_angles(path: Path | None) -> dict[str, dict[str, float]]:
    if path is None:
        return {}
    try:
        from openpyxl import load_workbook
    except Exception:
        return {}
    if not path.exists():
        return {}
    try:
        workbook = load_workbook(path, data_only=True)
    except Exception:
        return {}
    def _parse_sheet(rows: list[tuple[object, ...]], headers: list[str]) -> dict[str, dict[str, float]]:
        angles: dict[str, dict[str, float]] = {}
        header_lookup = {name.lower(): idx for idx, name in enumerate(headers) if name}
        angle_idx = header_lookup.get("anteversion_angle_deg")
        case_key_candidates = (
            "case_id",
            "case",
            "caseid",
            "case id",
            "patient_id",
            "patient id",
        )
        case_idx = None
        for key in case_key_candidates:
            idx = header_lookup.get(key)
            if idx is not None:
                case_idx = idx
                break
        if case_idx is None:
            for name, idx in header_lookup.items():
                if "case" in name and "id" in name:
                    case_idx = idx
                    break
        if angle_idx is not None and case_idx is not None:
            for row in rows[1:]:
                if case_idx >= len(row) or angle_idx >= len(row):
                    continue
                case_val = row[case_idx]
                angle_val = row[angle_idx]
                if case_val is None or angle_val is None:
                    continue
                case_id = _normalize_case_id(str(case_val))
                if not case_id:
                    continue
                try:
                    angle = float(angle_val)
                except (TypeError, ValueError):
                    continue
                angles.setdefault(case_id, {})["value"] = angle
            return angles
        # Fallback: case-per-row layout (Case in first column, values in remaining columns)
        if headers and headers[0].strip().lower() == "case":
            angle_columns = [
                idx
                for idx, header in enumerate(headers)
                if header and header.lower().startswith(("left_", "right_"))
            ]
            def _is_numeric(value: object) -> bool:
                try:
                    float(value)
                except (TypeError, ValueError):
                    return False
                return True
            for row in rows[1:]:
                if not row:
                    continue
                if angle_columns:
                    has_angle_value = False
                    valid_row = True
                    for idx in angle_columns:
                        if idx >= len(row):
                            continue
                        cell = row[idx]
                        if cell is None:
                            continue
                        has_angle_value = True
                        if not _is_numeric(cell):
                            valid_row = False
                            break
                    if not has_angle_value or not valid_row:
                        continue
                case_id = _normalize_case_id(str(row[0]))
                if not case_id:
                    continue
                for idx, cell in enumerate(row[1:], start=1):
                    if cell is None:
                        continue
                    header = headers[idx] if idx < len(headers) else ""
                    header_lower = header.lower()
                    if not header_lower:
                        continue
                    if header_lower.startswith("left_"):
                        side = "Left"
                        user = header[5:]
                    elif header_lower.startswith("right_"):
                        side = "Right"
                        user = header[6:]
                    else:
                        continue
                    user = user.strip().upper()
                    if not user:
                        continue
                    try:
                        angle = float(cell)
                    except (TypeError, ValueError):
                        continue
                    per_case = angles.setdefault(case_id, {})
                    key = f"{side}_{user}"
                    if key in per_case:
                        continue
                    per_case[key] = angle
            return angles
        # Fallback: look for a row keyed by anteversion_angle_deg
        if headers:
            for row in rows[1:]:
                row_key = str(row[0]).strip().lower() if row and row[0] is not None else ""
                if row_key != "anteversion_angle_deg":
                    continue
                for idx, header in enumerate(headers[1:], start=1):
                    case_id = _normalize_case_id(str(header))
                    if not case_id:
                        continue
                    if idx >= len(row):
                        continue
                    val = row[idx]
                    if val is None:
                        continue
                    try:
                        angle = float(val)
                    except (TypeError, ValueError):
                        continue
                    angles[case_id] = angle
                return angles
        return angles

    angles: dict[str, dict[str, float]] = {}
    primary_sheet = "Femoral_Anteversion_Angles"
    sheets_to_scan = [primary_sheet] if primary_sheet in workbook.sheetnames else list(workbook.sheetnames)
    if primary_sheet not in workbook.sheetnames:
        print("Anteversion sheet not found. Scanning all sheets.")
    for name in sheets_to_scan:
        sheet = workbook[name]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(value).strip() if value is not None else "" for value in rows[0]]
        angles = _parse_sheet(rows, headers)
        if angles:
            return angles
    if not angles:
        print("No anteversion values parsed; sample rows:")
        for name in sheets_to_scan[:1]:
            sheet = workbook[name]
            rows = list(sheet.iter_rows(values_only=True))
            for sample in rows[0:6]:
                print(f"  {sample}")
    return angles


def _populate_cases_sheet(
    sheet,
    rows: list[dict[str, Any]],
    zone_names: list[str],
    partition_zones: list[str],
    partition_tags: list[str],
    gruen_zone_ids: list[int],
    gruen_tags: list[str],
) -> None:
    headers = [
        "User",
        "Case",
        "Config Label",
        "Config Source",
        "Config Index",
        "Hip Config Name",
        "Hip Config Description",
        "Stem UID",
        "Stem Manufacturer",
        "Stem Enum",
        "Stem Friendly Name",
        "Rotation Mode",
        "Requested Side",
        "Configured Side",
        "Stem State",
        "Stem RCC ID",
        "Scalar Array",
        "Scalar Min",
        "Scalar Max",
        "Total Points",
        "Volume Name",
        "Volume Path",
        "Screenshot Folder",
        "Original Stem VTP",
        "Metrics XML",
        "Generated At",
    ]
    for zone_name in zone_names:
        headers.append(f"{zone_name} Count")
        headers.append(f"{zone_name} Percent (%)")
    for zone_name in partition_zones:
        for tag_name in partition_tags:
            headers.append(f"{zone_name} {tag_name} (%)")
    for zone_id in gruen_zone_ids:
        for tag_name in gruen_tags:
            headers.append(f"Gruen {zone_id} {tag_name} Count")
            headers.append(f"Gruen {zone_id} {tag_name} Percent (%)")
    sheet.append(headers)

    for entry in rows:
        stem = entry["stem"]
        volume_name = Path(entry["volume_path"]).name if entry["volume_path"] else ""
        config_index_value = entry.get("config_index")
        if config_index_value is None or config_index_value < 0:
            config_index_value = ""
        row = [
            entry["user_id"],
            entry["case_id"],
            entry.get("config_label", ""),
            entry.get("config_source", ""),
            config_index_value,
            entry.get("hip_config_name", ""),
            entry.get("hip_config_description", ""),
            stem.get("uid", ""),
            stem.get("manufacturer", ""),
            stem.get("enumName", ""),
            stem.get("friendlyName", ""),
            stem.get("rotationMode", ""),
            stem.get("requestedSide", ""),
            stem.get("configuredSide", ""),
            stem.get("state", ""),
            stem.get("rccId", ""),
            entry["array"],
            entry["scalar_min"],
            entry["scalar_max"],
            entry["total_points"],
            volume_name,
            entry["volume_path"],
            entry.get("screenshot_dir", ""),
            entry["original_vtp"],
            str(entry["path"]),
            entry["generated_at"],
        ]
        for zone_name in zone_names:
            zone = entry["zones"].get(zone_name)
            row.append(zone["count"] if zone else 0)
            row.append(zone["percent"] if zone else 0.0)
        partition_summary = entry.get("partition_summary") or {}
        for zone_name in partition_zones:
            zone_tags = partition_summary.get(zone_name, {})
            for tag_name in partition_tags:
                row.append(zone_tags.get(tag_name, 0.0))
        gruen_summary = entry.get("gruen_summary") or {}
        for zone_id in gruen_zone_ids:
            zone_tags = gruen_summary.get(zone_id, {})
            for tag_name in gruen_tags:
                count, percent = zone_tags.get(tag_name, (0, 0.0))
                row.append(count)
                row.append(percent)
        sheet.append(row)

    _autosize_columns(sheet)


def _populate_summary_sheet(sheet, rows: list[dict[str, Any]], zone_names: list[str]) -> None:
    sheet.title = "Users"
    headers = ["User", "Cases", "Configurations", "Total Points"]
    for zone_name in zone_names:
        headers.append(f"{zone_name} Count")
        headers.append(f"{zone_name} Percent (%)")
    sheet.append(headers)

    def _new_bucket():
        return {
            "configurations": 0,
            "case_ids": set(),
            "total_points": 0,
            "zones": {name: 0 for name in zone_names},
        }

    aggregates: dict[str, dict[str, Any]] = defaultdict(_new_bucket)

    for entry in rows:
        user = entry["user_id"] or "<unknown>"
        bucket = aggregates[user]
        bucket["configurations"] += 1
        case_id = entry.get("case_id")
        if case_id:
            bucket["case_ids"].add(case_id)
        bucket["total_points"] += entry["total_points"]
        for zone_name in zone_names:
            zone = entry["zones"].get(zone_name)
            if zone:
                bucket["zones"][zone_name] += zone["count"]

    for user in sorted(aggregates.keys()):
        bucket = aggregates[user]
        case_count = len(bucket["case_ids"])
        if case_count == 0:
            case_count = bucket["configurations"]
        row = [user, case_count, bucket["configurations"], bucket["total_points"]]
        for zone_name in zone_names:
            count = bucket["zones"][zone_name]
            total_points = bucket["total_points"]
            percent = (count / total_points * 100.0) if total_points else 0.0
            row.append(count)
            row.append(percent)
        sheet.append(row)

    _autosize_columns(sheet)


def _populate_gruen_sheet(
    sheet,
    rows: list[dict[str, Any]],
    gruen_zone_ids: list[int],
    gruen_tags: list[str],
    anteversion_angles: dict[str, dict[str, float]] | None = None,
    include_anteversion: bool = False,
) -> None:
    sheet.title = "GruenHU"
    anteversion_angles = anteversion_angles or {}
    headers = [
        "User",
        "Case",
        "Config Label",
        "Config Source",
        "Config Index",
        "Hip Config Name",
        "Hip Config Pretty Name",
        "Stem UID",
        "Requested Side",
        "Configured Side",
        "Metrics XML",
        "Gruen Zone",
        "HU Tag",
        "Count",
        "Percent (%)",
    ]
    if include_anteversion:
        headers.insert(7, "Anteversion Angle (deg)")
    sheet.append(headers)

    def _normalize_side(value: object) -> str:
        if value is None:
            return ""
        text = str(value).strip().lower()
        if text.startswith("l"):
            return "Left"
        if text.startswith("r"):
            return "Right"
        return ""

    def _infer_side_label(entry: dict[str, Any]) -> str:
        stem = entry.get("stem", {}) or {}
        side_label = _normalize_side(stem.get("configuredSide")) or _normalize_side(stem.get("requestedSide"))
        if side_label:
            return side_label
        candidates = (
            entry.get("config_label"),
            entry.get("hip_config_name"),
            entry.get("hip_config_description"),
            entry.get("config_source"),
        )
        for candidate in candidates:
            if not candidate:
                continue
            text = str(candidate).lower()
            if re.search(r"\bleft\b|\(l\)", text):
                return "Left"
            if re.search(r"\bright\b|\(r\)", text):
                return "Right"
        return ""

    for entry in rows:
        gruen_summary = entry.get("gruen_summary") or {}
        if not gruen_summary:
            continue
        config_index_value = entry.get("config_index")
        if config_index_value is None or config_index_value < 0:
            config_index_value = ""
        stem = entry.get("stem", {})
        for zone_id in gruen_zone_ids:
            zone_tags = gruen_summary.get(zone_id, {})
            for tag_name in gruen_tags:
                count, percent = zone_tags.get(tag_name, (0, 0.0))
                case_id_value = _normalize_case_id(str(entry.get("case_id", "")))
                if not case_id_value:
                    case_id_value = _extract_case_id_from_path(str(entry.get("path", "")))
                case_key = case_id_value
                side_label = _infer_side_label(entry)
                user_value = str(entry.get("user_id", "")).strip().upper()
                row = [
                    entry.get("user_id", ""),
                    case_id_value,
                    entry.get("config_label", ""),
                    entry.get("config_source", ""),
                    config_index_value,
                    entry.get("hip_config_name", ""),
                    entry.get("hip_config_description", ""),
                    stem.get("uid", ""),
                    stem.get("requestedSide", ""),
                    stem.get("configuredSide", ""),
                    str(entry.get("path", "")),
                    zone_id,
                    tag_name,
                    count,
                    percent,
                ]
                if include_anteversion:
                    anteversion_key = f"{side_label}_{user_value}" if side_label and user_value else ""
                    case_angles = anteversion_angles.get(case_key, {})
                    anteversion_value = case_angles.get(anteversion_key, "") if anteversion_key else ""
                    if anteversion_value == "" and user_value:
                        user_matches = [
                            val
                            for key, val in case_angles.items()
                            if key.endswith(f"_{user_value}")
                        ]
                        if len(user_matches) == 1:
                            anteversion_value = user_matches[0]
                    if anteversion_value == "" and side_label:
                        side_prefix = f"{side_label}_"
                        side_values = [
                            val for key, val in case_angles.items() if key.startswith(side_prefix)
                        ]
                        if side_values:
                            anteversion_value = sum(side_values) / len(side_values)
                    if anteversion_value == "" and "value" in case_angles:
                        anteversion_value = case_angles.get("value", "")
                    row.insert(7, anteversion_value)
                sheet.append(row)

    _autosize_columns(sheet)


def main() -> int:
    args = _parse_args()
    root = Path(args.root).expanduser().resolve()
    if not root.is_dir():
        print(f"Error: root directory '{root}' does not exist", file=sys.stderr)
        return 1
    output_path = Path(args.output).expanduser().resolve()
    anteversion_angles = _load_anteversion_angles(
        Path(args.anteversion_excel).expanduser().resolve() if args.anteversion_excel else None
    )
    if args.anteversion_excel:
        print("Anteversion angles loaded from Excel:")
        if not anteversion_angles:
            print("  (none found)")
        else:
            for case_id in sorted(anteversion_angles.keys()):
                case_angles = anteversion_angles[case_id]
                pairs = ", ".join(
                    f"{key}={case_angles[key]:.4f}" for key in sorted(case_angles.keys())
                )
                print(f"  {case_id}: {pairs}")

    xml_files = _collect_metric_files(root, args.pattern)
    if not xml_files:
        print("No *_stem_metrics.xml files found. Run load_nifti_and_stem with --compute-stem-scalars first.")
        return 0

    rows: list[dict[str, Any]] = []
    zone_names: set[str] = set()
    partition_zone_names: set[str] = set()
    partition_tag_names: set[str] = set()
    gruen_zone_ids: set[int] = set()
    gruen_tag_names: set[str] = set()
    for xml_file in xml_files:
        try:
            entry = _parse_metric_file(xml_file, args.hu_heatmap)
        except Exception as exc:
            print(f"Warning: failed to parse '{xml_file}': {exc}")
            continue
        rows.append(entry)
        zone_names.update(entry["zones"].keys())
        partition_summary = entry.get("partition_summary") or {}
        partition_zone_names.update(partition_summary.keys())
        for tag_map in partition_summary.values():
            partition_tag_names.update(tag_map.keys())
        gruen_summary = entry.get("gruen_summary") or {}
        gruen_zone_ids.update(gruen_summary.keys())
        for tag_map in gruen_summary.values():
            gruen_tag_names.update(tag_map.keys())
        if not args.quiet:
            print(f"Loaded metrics from {xml_file}")

    if not rows:
        print("No valid metrics parsed; nothing to export.")
        return 0

    zone_name_list = sorted(zone_names)
    partition_zone_list = sorted(partition_zone_names)
    partition_tag_list = sorted(partition_tag_names)
    gruen_zone_list = sorted(gruen_zone_ids)
    gruen_tag_list = sorted(gruen_tag_names)
    workbook = Workbook()
    cases_sheet = workbook.active
    cases_sheet.title = "Cases"
    _populate_cases_sheet(
        cases_sheet,
        rows,
        zone_name_list,
        partition_zone_list,
        partition_tag_list,
        gruen_zone_list,
        gruen_tag_list,
    )
    summary_sheet = workbook.create_sheet(title="Users")
    _populate_summary_sheet(summary_sheet, rows, zone_name_list)
    if gruen_zone_list and gruen_tag_list:
        gruen_sheet = workbook.create_sheet(title="GruenHU")
        _populate_gruen_sheet(
            gruen_sheet,
            rows,
            gruen_zone_list,
            gruen_tag_list,
            anteversion_angles,
            include_anteversion=bool(args.anteversion_excel),
        )
    else:
        if any(entry.get("gruen_xml_found") for entry in rows):
            gruen_sheet = workbook.create_sheet(title="GruenHU")
            _populate_gruen_sheet(
                gruen_sheet,
                rows,
                gruen_zone_list,
                gruen_tag_list,
                anteversion_angles,
                include_anteversion=bool(args.anteversion_excel),
            )
    workbook.save(output_path)
    print(f"Wrote Excel workbook: {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
